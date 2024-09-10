
#nltk.download('wordnet')
import nltk
import string
from openpyxl import load_workbook
import gensim
from sklearn.feature_extraction.text import CountVectorizer
from nltk.corpus import stopwords
from nltk.stem.porter import *
from nltk.stem import PorterStemmer as ps
import numpy as np
from nltk.stem import WordNetLemmatizer
from nltk import word_tokenize
from nltk.tokenize import SyllableTokenizer
from nltk.corpus import wordnet # The WordNet is a part of Python's Natural Language Toolkit
from sklearn.metrics import accuracy_score
import pandas as pd

###########################################################################################
stemmer = ps()  # for stemming the sentences of a user review
np.random.seed(2018)


wb = load_workbook('Classification.xlsx', read_only=True)  # loading the excel file
wss = wb.get_sheet_by_name('Security Class')  # selecting the sheet

###############################################################################################
use_col = 1
x2 = np.array([r[use_col].value for r in wss.iter_rows()])  # iterating the rows in a sheet
res = []
for val in x2:
    if val != None:
        res.append(val)



def lemmatize_stemming(text):
    return stemmer.stem(WordNetLemmatizer().lemmatize(text, pos='v'))

def preprocess(text):
    result = []
    for token in gensim.utils.simple_preprocess(text):
        if token not in gensim.parsing.preprocessing.STOPWORDS and len(token) > 3:
            result.append(lemmatize_stemming(token))
    return result


###############################################################################################
# cleaning the data
doc_sample = res[1:]  # removing the title and considering the remaining user review
words = []
for word in doc_sample:
    tk = SyllableTokenizer()
    text_tokens = tk.tokenize(word)
    tokens_without_sw = [word for word in text_tokens if not word in stopwords.words()]
    words.append(tokens_without_sw)

stopwords = nltk.corpus.stopwords.words('english')  # this will remove stop words already present in nltkcorpus
wn = nltk.WordNetLemmatizer()


def clean_text(text):  # this will clean the data as well as split/tokenized it and find the sysnonmn of each word
    text = "".join([word for word in text if word not in string.punctuation])
    tokens = re.split("\W+", text)
    text = [wn.lemmatize(word) for word in tokens if word not in stopwords]
    lemmas = []
    for token in text:
        lemmas += [synset.lemmas()[0].name() for synset in
                   wordnet.synsets(token)]  # this will find the sysnomum of a word
    return list(set(lemmas))


# print(clean_text(res))
########################################################################################
# Augmenting the user review
listforAug = []
AugmentView = []
res = res[1:]
for i in words:
    words = (" ").join(i)
    listforAug.append(words)

for i in listforAug:
    a = clean_text(i)
    a = (" ").join(a)
    AugmentView.append(str(i) + str(a))


d = []
for i in AugmentView:
    print("Final List ", str(i))
    d.append(i)

###########################################################################################
##correct
# tokenizing the Augmented user reviews and again check the stopwords
from nltk.corpus import stopwords

finallist = []
for word in AugmentView:
    finallist.append(word_tokenize(word))
for i in range(len(finallist)):
    finallist[i] = [w for w in finallist[i] if w not in stopwords.words('english')]



##########################################################################################
# Bag of WOrd to find the weight of given word
def word_extraction(sentence):
    ignore_words = ['a']
    words = re.sub("[^w]", " ", sentence).split()  # nltk.word_tokenize(sentence)
    words_cleaned = [w.lower() for w in words if w not in ignore_words]
    return words_cleaned


def tokenize(sentences):
    words = []
    for sentence in sentences:
        w = word_extraction(sentence)
        words.extend(w)
        words = sorted(list(set(words)))
        return words


def generate_bow(allsentences):
    vocab = tokenize(allsentences)
    print("Word List for Document \n{0} \n".format(vocab))
    for sentence in allsentences:
        words = word_extraction(sentence)
        bag_vector = np.zeros(len(vocab))
        for w in words:
            for i, word in enumerate(vocab):
                if word == w:
                    bag_vector[i] += 1


# generate_bow(AugmentView)
###################################################################################

vectorizer = CountVectorizer(ngram_range=(1, 1), stop_words='english')
X = vectorizer.fit_transform(res)
cv_dataframe = pd.DataFrame(X.toarray(), columns=vectorizer.get_feature_names())


dictionaryObject={}
dictionaryObject = cv_dataframe.to_dict(orient="list")


# Updated class

line = []
l1= 20*('Security',)
l2= 119*('Portability' ,)
l3= 120*('Performance Efficency',)
l4= 587*('Relaibility',)
l5= 432*('Usability',)
functional=l1+l2+l3+l4+l5

dictionaryObject['Functional']=functional
list1 = []
for key in dictionaryObject.keys():
    list1.append(key)



train = pd.DataFrame.from_dict(dictionaryObject, orient='index')
train=train.transpose()
net = []
Lst =list1
for first, second in zip(Lst, Lst[1:]):
    net.append((first, second))

print("\n***************** Splitting the Data *************************\n")
msk = np.random.rand(len(train)) < 0.7
train = train[msk]
test= train[~msk]

from pgmpy.models import BayesianModel

print("\n*****************Model Fitting*************************\n")

model = BayesianModel(net)  # conncted Grapgh
model.fit(train)

print("\n*****************Model Predicting *************************\n")
predict_data = test.copy()
predict_data.drop('Functional', axis=1, inplace=True)
y_pred = model.predict(predict_data)
y_actual=test[['Functional']]

print("\n*****************Accuracy of the Model*************************\n")
Acc= accuracy_score(y_actual, y_pred)
print(Acc)  # 55 %
