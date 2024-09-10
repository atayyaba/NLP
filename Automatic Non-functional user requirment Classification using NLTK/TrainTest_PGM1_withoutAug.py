import string
from collections import defaultdict
from openpyxl import Workbook
from gensim.models import Word2Vec
from openpyxl import load_workbook
from pandas import np
import gensim
from sklearn.feature_extraction.text import CountVectorizer
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize, util
from nltk.tokenize import sent_tokenize, word_tokenize
from gensim.utils import simple_preprocess
from gensim.parsing.preprocessing import STOPWORDS
from nltk.stem import WordNetLemmatizer, SnowballStemmer
from nltk.stem.porter import *
from nltk.stem import PorterStemmer as ps
import numpy as n
from nltk.corpus import wordnet
from jaccard_index.jaccard import jaccard_index
from sklearn.metrics import accuracy_score

from BoW import wordnet_lemmatizer
from nltk.corpus import wordnet  # The WordNet is a part of Python's Natural Language Toolkit
import networkx as nx
import matplotlib.pyplot as plt
import pylab
import pandas as pd

from nltk.corpus import wordnet # The WordNet is a part of Python's Natural Language Toolkit
import networkx as nx
import pylab

###########################################################################################
stemmer = ps()  # for stemming the sentences of a user review
n.random.seed(2018)
import nltk

nltk.download('wordnet')

wb = load_workbook('Classification.xlsx', read_only=True)  # loading the excel file
print(wb.sheetnames)
wss = wb.get_sheet_by_name('Security Class')  # selecting the sheet

###############################################################################################
use_col = 1
x2 = np.array([r[use_col].value for r in wss.iter_rows()])  # iterating the rows in a sheet
res = []
for val in x2:
    if val != None:
        res.append(val)
print(res)
print(len(res))
print("Original Document : ", res[1:])


def lemmatize_stemming(text):
    return stemmer.stem(WordNetLemmatizer().lemmatize(text, pos='v'))


def preprocess(text):
    result = []
    for token in gensim.utils.simple_preprocess(text):
        if token not in gensim.parsing.preprocessing.STOPWORDS and len(token) > 3:
            result.append(lemmatize_stemming(token))
    return result


###############################################################################################
# cleaning the data
doc_sample = res[1:]  # removing the title and considering the remaining user review
print('Tokenized document: ')
words = []
for word in doc_sample:
    text_tokens = word_tokenize(word)
    tokens_without_sw = [word for word in text_tokens if not word in stopwords.words()]
    words.append(tokens_without_sw)
print(words)

stopwords = nltk.corpus.stopwords.words('english')  # this will remove stop words already present in nltkcorpus
wn = nltk.WordNetLemmatizer()


def clean_text(text):  # this will clean the data as well as split/tokenized it and find the sysnonmn of each word
    text = "".join([word for word in text if word not in string.punctuation])
    tokens = re.split("\W+", text)
    text = [wn.lemmatize(word) for word in tokens if word not in stopwords]
    lemmas = []
    for token in text:
        lemmas += [synset.lemmas()[0].name() for synset in
                   wordnet.synsets(token)]  # this will find the sysnomum of a word
    return list(set(lemmas))


# print(clean_text(res))
########################################################################################
# Augmenting the user review
listforAug = []
AugmentView = []
res = res[1:]
for i in words:
    words = (" ").join(i)
    listforAug.append(words)
print(listforAug)
for i in listforAug:
    print("Without Augment: ", i)
    a = clean_text(i)
    a = (" ").join(a)
    print(" Synonms: ", a)
    AugmentView.append(str(i) + str(a))
    print(" Augment: ", str(AugmentView))

d = []
for i in AugmentView:
    print("Final List ", str(i))
    d.append(i)
"""    
f = open('Aug.txt', 'w', encoding='utf-8')
for i in d:
    f.write(i + "\n")
f.close()"""
###########################################################################################
##correct
# tokenizing the Augmented user reviews and again check the stopwords
from nltk.corpus import stopwords

finallist = []
for word in AugmentView:
    finallist.append(word_tokenize(word))
print("List before Vector: ", finallist)
for i in range(len(finallist)):
    finallist[i] = [w for w in finallist[i] if w not in stopwords.words('english')]
print("List before Vector Stop words: ", finallist)


##########################################################################################
# Bag of WOrd to find the weight of given word
def word_extraction(sentence):
    ignore_words = ['a']
    words = re.sub("[^w]", " ", sentence).split()  # nltk.word_tokenize(sentence)
    words_cleaned = [w.lower() for w in words if w not in ignore_words]
    return words_cleaned


def tokenize(sentences):
    words = []
    for sentence in sentences:
        w = word_extraction(sentence)
        words.extend(w)
        words = sorted(list(set(words)))
        return words


def generate_bow(allsentences):
    vocab = tokenize(allsentences)
    print("Word List for Document \n{0} \n".format(vocab))
    for sentence in allsentences:
        words = word_extraction(sentence)
        bag_vector = np.zeros(len(vocab))
        for w in words:
            for i, word in enumerate(vocab):
                if word == w:
                    bag_vector[i] += 1
                    print("{0}\n{1}\n".format(sentence, np.array(bag_vector)))


# generate_bow(AugmentView)
###################################################################################
vectorizer = CountVectorizer(ngram_range=(1, 1), stop_words='english')
X = vectorizer.fit_transform(res)
cv_dataframe = pd.DataFrame(X.toarray(), columns=vectorizer.get_feature_names())
print(cv_dataframe)

dictionaryObject={}
dictionaryObject = cv_dataframe.to_dict(orient="list")

print("DataFrame as a dictionary:")

print(dictionaryObject)
print("data d: ", dictionaryObject.keys())
print(len(dictionaryObject))



G=nx.DiGraph()
pylab.figure(1,figsize=(10,10)) # Size of the Grapgh Figure

line = []
l1= 20*('Security',)
l2= 119*('Portability' ,)
l3= 120*('Performance Efficency',)
l4= 587*('Relaibility',)
l5= 432*('Usability',)
functional=l1+l2+l3+l4+l5
print(len(functional))
dictionaryObject['Functional']=functional
list1 = []
for key in dictionaryObject.keys():
    list1.append(key)
print(len(list1))

#train= pd.DataFrame(dictionaryObject)
train = pd.DataFrame.from_dict(dictionaryObject, orient='index')
#train.transpose()
train=train.transpose()
#train = pd.DataFrame(dictionaryObject, columns=list)


list2 = list1
prob = []  # to store relationship between two pair of words
node = []  # to store the pairs it self in the form of list

temp = []

# find the relationship between words by finding the samentic similirity
for word1 in list1:
    for word2 in list2:

        # cognitive synonyms of the words are called synsets
        wordFromList1 = wordnet.synsets(word1)
        wordFromList2 = wordnet.synsets(word2)
        if wordFromList1 and wordFromList2:
            s = wordFromList1[0].wup_similarity(wordFromList2[0])
            if (word1 != word2):  # restrict comperison of word to it self
                if (s is None):  # of there are no cognitive relationship then ignore it
                    break;
                temp.append(word1)
                temp.append(word2)  # Create a temporary list
                prob.append(s)  # save the relationship / probability of similarity in weight class
                node.append((word1, word2))  # Add both words to Node List
similarity = ['%.2f' % elem for elem in
              prob]  # Round off the value of the relationship between words

# __________________________________________________________________________________________

# create Grapgh of the words and their similarity

i = -1
for data in node:
    for j in data:
        G.add_node(j)  # add words
    i = i + 1
    G.add_edge(data[0], data[1],
               weight=similarity[i])  # add edges between pair of words and shows relationship as their weight

# Shortant the Graph

for node in G.nodes():
    minimum = []
    edges = G.in_edges(node, data=True)
    if len(edges) > 0:
        for edge in edges:
            minimum.append(edge[2]['weight'])
        max_weight = max(minimum)
        for edge in list(edges):
            v = edge[2]['weight']
            if v < max_weight:
                print(edge[0], edge[1])
                G.remove_edge(edge[0], edge[1])

# __________________________________________________________________________________________
# Remove Cycles and duplicat edges to make the network for PGM


import networkx as nx

s=list(nx.simple_cycles(G))

for i in range(0,len(s)):
    y=s[i]
    target1=y[0]
    target2=y[1]
    for node in G.nodes():
        edges = G.in_edges(node, data=True)
        for edge in list(edges):
            if(target1==edge[0] and target2==edge[1]):
                print(edge[0], edge[1])
                G.remove_edge(edge[0], edge[1])
net=list(G.edges())
#_____________________________________________________________________________

Data = pd.DataFrame.from_dict(dictionaryObject, orient='index')
Data=Data.transpose()

print("\n***************** Splitting the Data *************************\n")
msk = np.random.rand(len(Data)) < 0.7
train = Data[msk]
test= Data[~msk]

from pgmpy.models import BayesianModel

print("\n*****************Model Fitting*************************\n")

model = BayesianModel(net)  # conncted Grapgh
model.fit(train)

print("\n*****************Model Predicting *************************\n")
predict_data = test.copy()
predict_data.drop('Functional', axis=1, inplace=True)
y_pred = model.predict(predict_data)
y_actual=test[['Functional']]

print("\n*****************Accuracy of the Model*************************\n")
Acc= accuracy_score(y_actual, y_pred)
print(Acc)  # 52 %


